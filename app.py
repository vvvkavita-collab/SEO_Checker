import streamlit as st
import pandas as pd
import requests
from bs4 import BeautifulSoup
from urllib.parse import urlparse
from io import BytesIO
import re
import json
import unicodedata
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ================= PAGE CONFIG =================
st.set_page_config(page_title="Advanced SEO Auditor – Google Guidelines", layout="wide")
st.title("🧠 Advanced SEO Auditor")

# ================= CSS =================
st.markdown("""
<style>
div[data-testid="stDataFrame"] table th {
    text-align: center !important;
}
div[data-testid="stDataFrame"] table td {
    vertical-align: middle;
}
div[data-testid="stDataFrame"] table td:nth-child(3),
div[data-testid="stDataFrame"] table td:nth-child(4),
div[data-testid="stDataFrame"] table td:nth-child(5) {
    text-align: center !important;
}
</style>
""", unsafe_allow_html=True)

HEADERS = {"User-Agent": "Mozilla/5.0"}

# ================= SIDEBAR =================
st.sidebar.header("SEO Mode")
bulk_file = st.sidebar.file_uploader("Upload Bulk URLs (TXT / CSV)", type=["txt", "csv"])
url_input = st.text_input("Paste URL")
analyze = st.button("Analyze")

# ================= STOP WORDS =================
TITLE_STOP_WORDS = ["breaking","exclusive","shocking","must read","update","alert","latest","big","viral"]

URL_STOP_WORDS = [
    "for","today","latest","news","update","information","details","story","article",
    "about","on","in","to","of","with",
    "current","recent","new",
    "breaking","exclusive","viral","shocking","must","read",
    "what","why","how","when","where","who",
    "page","pages","index","view","print","amp","category","tag"
]

# ================= HELPERS =================
def get_soup(url):
    r = requests.get(url, headers=HEADERS, timeout=25, verify=False)
    r.raise_for_status()
    return BeautifulSoup(r.text, "lxml")

def get_article(soup):
    return soup.find("article") or soup

def visible_len(text):
    return sum(1 for c in text if not unicodedata.category(c).startswith("C"))

def safe_text(el):
    return el.get_text(" ", strip=True) if el else ""

def get_real_paragraphs(article):
    return [p.get_text(" ", strip=True) for p in article.find_all("p") if len(p.get_text(strip=True)) > 80]

def get_links(article, domain):
    internal = external = 0
    for a in article.find_all("a", href=True):
        if domain in a["href"]:
            internal += 1
        else:
            external += 1
    return internal, external

def extract_meta_image(soup):
    og = soup.find("meta", property="og:image")
    return og["content"] if og else None

def has_schema(soup):
    return bool(soup.find("script", type="application/ld+json"))

# ================= EXCEL FORMAT =================
def format_excel(sheets):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for name, df in sheets.items():
            df.to_excel(writer, sheet_name=name, index=False)

    output.seek(0)
    wb = load_workbook(output)

    for ws in wb.worksheets:
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        bold = Font(bold=True)
        border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        for col in ws.columns:
            ws.column_dimensions[get_column_letter(col[0].column)].width = 30

        for cell in ws[1]:
            cell.font = bold
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    final = BytesIO()
    wb.save(final)
    final.seek(0)
    return final

# ================= ANALYSIS =================
def analyze_url(url):
    soup = get_soup(url)
    article = get_article(soup)
    domain = urlparse(url).netloc

    title = safe_text(soup.find("h1") or soup.find("title"))
    title_len = visible_len(title)

    paras = get_real_paragraphs(article)
    word_count = sum(len(p.split()) for p in paras)

    internal, external = get_links(article, domain)
    meta_img = extract_meta_image(soup)
    schema = has_schema(soup)

    score = 100
    if title_len < 55 or title_len > 70: score -= 12
    if word_count < 300: score -= 12
    if internal < 2: score -= 5
    if not meta_img: score -= 5
    if not schema: score -= 10

    audit_df = pd.DataFrame([
        ["Title Length", title_len, "55–70", "⚠️" if title_len < 55 or title_len > 70 else "✅"],
        ["Word Count", word_count, "300+", "⚠️" if word_count < 300 else "✅"],
        ["Internal Links", internal, "2+", "⚠️" if internal < 2 else "✅"],
        ["Meta Image", "Yes" if meta_img else "No", "Yes", "⚠️" if not meta_img else "✅"],
        ["Schema", "Yes" if schema else "No", "Yes", "⚠️" if not schema else "✅"],
        ["Final SEO Score", f"{score}/100", "≥80", "⚠️" if score < 80 else "✅"],
    ], columns=["Metric","Actual","Ideal","Verdict"])

    grading_df = pd.DataFrame([
        ["Base Score", 100],
        ["Title Issue", -12 if title_len < 55 or title_len > 70 else 0],
        ["Low Content", -12 if word_count < 300 else 0],
        ["Low Internal Links", -5 if internal < 2 else 0],
        ["No Meta Image", -5 if not meta_img else 0],
        ["No Schema", -10 if not schema else 0],
        ["Final Score", score]
    ], columns=["Rule","Value"])

    return audit_df, grading_df

# ================= RUN =================
urls = set()
if bulk_file:
    raw = bulk_file.read().decode("utf-8", errors="ignore")
    urls.update([l.strip() for l in raw.splitlines() if l.strip()])

if url_input:
    urls.add(url_input.strip())

if analyze and urls:
    audits, scores = [], []

    for u in urls:
        st.subheader(f"📊 SEO Audit – {u}")
        audit_df, grading_df = analyze_url(u)
        st.dataframe(audit_df, use_container_width=True)
        st.dataframe(grading_df, use_container_width=False)

        audit_df.insert(0, "URL", u)
        grading_df.insert(0, "URL", u)
        audits.append(audit_df)
        scores.append(grading_df)

    excel_file = format_excel({
        "SEO Audit": pd.concat(audits, ignore_index=True),
        "Score Logic": pd.concat(scores, ignore_index=True)
    })

    st.download_button(
        "⬇️ Download SEO Audit Excel",
        data=excel_file,
        file_name="SEO_Audit_Report.xlsx"
    )
